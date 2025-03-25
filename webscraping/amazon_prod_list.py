from selenium import webdriver
import openpyxl
import time
from selenium.webdriver.support.wait import WebDriverWait
from selenium import webdriver
from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options



print("Process Started...")

# This will start the Chrome driver in headless mode, which means that it will run without opening a visible browser window.
# options = Options()
# options.add_argument('--headless')
# driver = webdriver.Chrome(options=options)

# create a new Chrome web driver
driver = webdriver.Chrome()

# navigate to the URL of the page we want to scrape
driver.get("https://www.amazon.com/s?k=iphone")
# driver.get("https://www.amazon.com/s?k=redmi&crid=1JBVNMFPCR5VW&sprefix=redmi%2Caps%2C303&ref=nb_sb_noss_1")
# driver.get("https://www.amazon.com/s?k=laptop")

# find the product titles on the page in square bracket we can increrase product number
# product_titles = driver.find_elements_by_xpath("//span[@class='.a-size-medium.a-color-base.a-text-normal']")[:2]
product_titles = driver.find_elements(By.CSS_SELECTOR, ".a-size-medium.a-color-base.a-text-normal")[:20]
time.sleep(10)   
# extract the text content of the product titles
product_titles = [title.text for title in product_titles]

# create a new Excel workbook and add a worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# add the product titles to the worksheet
worksheet.cell(row=1, column=1, value="Product Name")
for i, title in enumerate(product_titles):
    worksheet.cell(row=i+2, column=1, value=title)

# save the workbook
# workbook.save("amazon_laptops.xlsx")
workbook.save("amazon_iphone.xlsx")

# close the web driver
driver.quit()
print("\n********Process complete File is Ready.*********\n")
